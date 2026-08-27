"""账单解析：微信 xlsx / 支付宝 CSV（GBK）。

统一输出 dict 列表，字段与 transactions 表对应（不含 category_id）：
    order_no, tx_time, counterparty, description, amount, refund_amount,
    pay_method, alipay_category
"""
import csv
import io
import re
from datetime import datetime

from openpyxl import load_workbook


def _parse_time(value):
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    s = str(value).strip()
    for fmt in ('%Y/%m/%d %H:%M:%S', '%Y/%m/%d %H:%M',
                '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d %H:%M:%S')
        except ValueError:
            continue
    raise ValueError(f'无法解析交易时间: {s!r}')


def _parse_refund(status, amount):
    """微信账单的退款信息标在原交易行的“当前状态”列上。"""
    if not status:
        return 0.0
    if '已全额退款' in status:
        return amount
    m = re.search(r'已退款[（(]¥?([\d.]+)[)）]', str(status))
    return float(m.group(1)) if m else 0.0


def parse_wechat(file_or_path):
    """微信支付账单流水 xlsx。列：交易时间/交易类型/交易对方/商品/收/支/金额(元)/
    支付方式/当前状态/交易单号/商户单号/备注"""
    ws = load_workbook(file_or_path, read_only=True).active
    rows = [r for r in ws.iter_rows(values_only=True)]
    # 定位表头行（有的导出带说明行前缀）
    header_idx = None
    for i, r in enumerate(rows[:30]):
        if r and any(str(c).strip() == '交易时间' for c in r if c is not None):
            header_idx = i
            break
    if header_idx is None:
        raise ValueError('未找到微信账单表头（需包含“交易时间”列）')
    col = {str(c).strip(): i for i, c in enumerate(rows[header_idx]) if c is not None}

    out = []
    for r in rows[header_idx + 1:]:
        if not r or not r[col['交易时间']]:
            continue
        if str(r[col['收/支']]).strip() != '支出':
            continue  # 只记支出
        amount = float(str(r[col['金额(元)']]).replace(',', ''))
        refund = _parse_refund(r[col['当前状态']], amount)
        out.append({
            'order_no': str(r[col['交易单号']]).strip(),
            'tx_time': _parse_time(r[col['交易时间']]),
            'counterparty': str(r[col['交易对方']] or '').strip(),
            'description': str(r[col['商品']] or '').strip(),
            'amount': round(amount - refund, 2),
            'refund_amount': refund,
            'pay_method': str(r[col['支付方式']] or '').strip(),
            'alipay_category': '',
        })
    return out


def parse_alipay(file_or_stream):
    """支付宝交易明细 CSV（GBK 编码）。列：交易时间/交易分类/交易对方/对方账号/
    商品说明/收/支/金额/收/付款方式/交易状态/交易订单号/商家订单号/备注"""
    if hasattr(file_or_stream, 'read'):
        raw = file_or_stream.read()
    else:
        with open(file_or_stream, 'rb') as f:
            raw = f.read()
    try:
        text = raw.decode('gbk')
    except UnicodeDecodeError:
        text = raw.decode('utf-8-sig')

    rows = list(csv.reader(io.StringIO(text)))
    header_idx = None
    for i, r in enumerate(rows[:30]):
        if r and str(r[0]).strip() == '交易时间':
            header_idx = i
            break
    if header_idx is None:
        raise ValueError('未找到支付宝账单表头（需包含“交易时间”列）')
    col = {c.strip(): i for i, c in enumerate(rows[header_idx]) if c and c.strip()}

    out = []
    for r in rows[header_idx + 1:]:
        if len(r) < 12 or not r[col['交易时间']].strip():
            continue
        if r[col['收/支']].strip() != '支出':
            continue
        status = r[col['交易状态']].strip()
        if '关闭' in status or '失败' in status:
            continue  # 钱没扣成功，不导入
        out.append({
            'order_no': r[col['交易订单号']].strip(),
            'tx_time': _parse_time(r[col['交易时间']]),
            'counterparty': r[col['交易对方']].strip(),
            'description': r[col['商品说明']].strip(),
            'amount': float(r[col['金额']].replace(',', '')),
            'refund_amount': 0.0,
            'pay_method': r[col['收/付款方式']].strip(),
            'alipay_category': r[col['交易分类']].strip(),
        })
    return out
